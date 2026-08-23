"""
Generate a 10,000-person synthetic Shanghai population.

This script reproduces the population-generation logic used for the provided
CSV/XLSX dataset. Geographic fields are intentionally left blank for a later
map-binding stage.

Outputs:
    shanghai_synthetic_population_10000.csv
    shanghai_synthetic_population_10000.xlsx

Dependencies:
    Python 3.10+
    openpyxl
"""

from __future__ import annotations

import csv
import random
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

SEED = 20260819
N = 10_000


def weighted_choice(items):
    r = random.random()
    cumulative = 0.0
    for value, probability in items:
        cumulative += probability
        if r <= cumulative:
            return value
    return items[-1][0]


def exact_categories(items, n):
    result = []
    remaining = n
    for i, (category, probability) in enumerate(items):
        count = remaining if i == len(items) - 1 else round(n * probability)
        remaining -= count
        result.extend([category] * count)
    random.shuffle(result)
    return result[:n]


GENDER_PROBS = [("男", 0.52), ("女", 0.48)]
AGE_GROUP_PROBS = [
    ("18-24", 0.15),
    ("25-29", 0.20),
    ("30-34", 0.23),
    ("35-39", 0.19),
    ("40-44", 0.14),
    ("45-49", 0.09),
]
HUKOU_TYPE_PROBS = [("上海户籍", 0.45), ("外省市户籍", 0.55)]

AGE_RANGES = {
    "18-24": (18, 24),
    "25-29": (25, 29),
    "30-34": (30, 34),
    "35-39": (35, 39),
    "40-44": (40, 44),
    "45-49": (45, 49),
}

FAMILY_PROBS = {
    "18-24": [("无需要照顾的孩子", 0.94), ("有未成年孩子", 0.06), ("有成年孩子", 0.00)],
    "25-29": [("无需要照顾的孩子", 0.72), ("有未成年孩子", 0.28), ("有成年孩子", 0.00)],
    "30-34": [("无需要照顾的孩子", 0.43), ("有未成年孩子", 0.56), ("有成年孩子", 0.01)],
    "35-39": [("无需要照顾的孩子", 0.25), ("有未成年孩子", 0.70), ("有成年孩子", 0.05)],
    "40-44": [("无需要照顾的孩子", 0.20), ("有未成年孩子", 0.62), ("有成年孩子", 0.18)],
    "45-49": [("无需要照顾的孩子", 0.18), ("有未成年孩子", 0.42), ("有成年孩子", 0.40)],
}

OCCUPATION_BY_AGE = {
    "18-24": [("office_worker", 0.28), ("service_worker", 0.23), ("manual_worker", 0.11), ("freelancer", 0.08), ("university_student", 0.30)],
    "25-29": [("office_worker", 0.50), ("service_worker", 0.22), ("manual_worker", 0.11), ("freelancer", 0.10), ("university_student", 0.07)],
    "30-34": [("office_worker", 0.52), ("service_worker", 0.23), ("manual_worker", 0.12), ("freelancer", 0.11), ("university_student", 0.02)],
    "35-39": [("office_worker", 0.49), ("service_worker", 0.25), ("manual_worker", 0.14), ("freelancer", 0.12), ("university_student", 0.00)],
    "40-44": [("office_worker", 0.43), ("service_worker", 0.28), ("manual_worker", 0.17), ("freelancer", 0.12), ("university_student", 0.00)],
    "45-49": [("office_worker", 0.36), ("service_worker", 0.31), ("manual_worker", 0.21), ("freelancer", 0.12), ("university_student", 0.00)],
}

OCCUPATION_CN = {
    "office_worker": "上班族",
    "service_worker": "服务业者",
    "manual_worker": "体力劳动者",
    "freelancer": "自由业者",
    "university_student": "大学生",
}

NON_SHANGHAI_PROVINCES = [
    ("江苏", 0.18), ("安徽", 0.16), ("浙江", 0.11), ("河南", 0.10),
    ("山东", 0.07), ("江西", 0.06), ("湖北", 0.06), ("四川", 0.06),
    ("湖南", 0.05), ("福建", 0.03), ("河北", 0.03), ("广东", 0.02),
    ("陕西", 0.02), ("其他", 0.05),
]

SURNAMES = [
    ("王", .08), ("李", .074), ("张", .07), ("刘", .053), ("陈", .049),
    ("杨", .032), ("黄", .03), ("赵", .028), ("周", .026), ("吴", .024),
    ("徐", .02), ("孙", .018), ("胡", .017), ("朱", .016), ("高", .015),
    ("林", .014), ("何", .013), ("郭", .013), ("马", .012), ("罗", .012),
    ("梁", .011), ("宋", .010), ("郑", .010), ("谢", .009), ("韩", .009),
    ("唐", .008), ("冯", .008), ("于", .007), ("董", .007), ("程", .006),
    ("曹", .006), ("袁", .006), ("邓", .006), ("许", .006),
]
_surname_total = sum(p for _, p in SURNAMES)
SURNAMES = [(name, p / _surname_total) for name, p in SURNAMES]

MALE_CHARS = list("宇轩浩然子涵俊杰嘉豪明哲博文志远天佑泽楷晨阳睿哲文昊嘉诚景轩梓豪")
FEMALE_CHARS = list("欣怡雨桐诗涵婉婷佳琪思涵梦瑶雅婷子晴若曦语嫣静怡佳宁晓彤可欣")
NEUTRAL_CHARS = list("晨宁安乐嘉言清羽知远一诺思源可心星辰")

SERVICE_TITLES = ["餐厅服务员", "厨师", "店员", "收银员", "酒店前台", "酒店服务人员", "美容师", "理发师", "健身教练", "房产经纪", "客服", "保安"]
MANUAL_TITLES = ["建筑工人", "电工", "焊工", "维修技师", "仓库操作员", "搬运工", "快递员", "配送员", "司机", "设备操作员", "安装工"]
FREELANCER_TITLES = ["自由设计师", "摄影师", "自媒体从业者", "独立程序员", "咨询顾问", "翻译", "私教", "网店经营者", "自由销售"]
OFFICE_JUNIOR = ["行政专员", "财务专员", "会计", "销售代表", "运营专员", "软件工程师", "设计师", "数据分析师", "采购专员", "客户专员"]
OFFICE_MID = ["产品经理", "项目经理", "客户经理", "高级软件工程师", "高级运营专员", "财务主管", "销售经理", "HR经理", "数据分析师", "采购经理"]
OFFICE_SENIOR = ["部门经理", "区域经理", "高级项目经理", "高级产品经理", "技术经理", "运营经理", "财务经理", "销售经理", "人力资源经理", "业务负责人"]


def make_name(gender):
    surname = weighted_choice(SURNAMES)
    pool = MALE_CHARS if gender == "男" else FEMALE_CHARS
    if random.random() < 0.22:
        return surname + random.choice(pool + NEUTRAL_CHARS)
    return surname + random.choice(pool + NEUTRAL_CHARS) + random.choice(pool + NEUTRAL_CHARS)


def choose_title(occupation, age):
    if occupation == "office_worker":
        if age <= 24:
            tier = weighted_choice([("junior", .94), ("mid", .06), ("senior", 0)])
        elif age <= 29:
            tier = weighted_choice([("junior", .72), ("mid", .27), ("senior", .01)])
        elif age <= 34:
            tier = weighted_choice([("junior", .45), ("mid", .50), ("senior", .05)])
        elif age <= 39:
            tier = weighted_choice([("junior", .28), ("mid", .58), ("senior", .14)])
        elif age <= 44:
            tier = weighted_choice([("junior", .20), ("mid", .55), ("senior", .25)])
        else:
            tier = weighted_choice([("junior", .18), ("mid", .50), ("senior", .32)])
        return random.choice({"junior": OFFICE_JUNIOR, "mid": OFFICE_MID, "senior": OFFICE_SENIOR}[tier])

    if occupation == "university_student":
        if age <= 22:
            return weighted_choice([("本科生", .86), ("硕士研究生", .13), ("博士研究生", .01)])
        if age <= 25:
            return weighted_choice([("本科生", .25), ("硕士研究生", .65), ("博士研究生", .10)])
        return weighted_choice([("本科生", .05), ("硕士研究生", .55), ("博士研究生", .40)])

    if occupation == "service_worker":
        return random.choice(SERVICE_TITLES)
    if occupation == "manual_worker":
        return random.choice(MANUAL_TITLES)
    return random.choice(FREELANCER_TITLES)


def generate_rows(n=N, seed=SEED):
    random.seed(seed)
    genders = exact_categories(GENDER_PROBS, n)
    age_groups = exact_categories(AGE_GROUP_PROBS, n)
    hukou_types = exact_categories(HUKOU_TYPE_PROBS, n)

    rows = []
    for i in range(n):
        gender = genders[i]
        age_group = age_groups[i]
        age = random.randint(*AGE_RANGES[age_group])
        family = weighted_choice(FAMILY_PROBS[age_group])
        occupation = weighted_choice(OCCUPATION_BY_AGE[age_group])
        hukou_type = hukou_types[i]
        province = "上海" if hukou_type == "上海户籍" else weighted_choice(NON_SHANGHAI_PROVINCES)

        rows.append([
            f"P{i + 1:05d}",
            make_name(gender),
            gender,
            age,
            age_group,
            family,
            OCCUPATION_CN[occupation],
            occupation,
            choose_title(occupation, age),
            hukou_type,
            province,
            "",  # home_place_id
            "",  # work_place_id
            "",  # school_place_id
        ])
    return rows


HEADERS = [
    "person_id", "姓名", "性别", "年龄", "年龄段", "家庭状态", "职业大类",
    "occupation_code", "具体职位", "户口类型", "户口省份",
    "home_place_id", "work_place_id", "school_place_id",
]


def write_csv(rows, path):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(HEADERS)
        writer.writerows(rows)


def write_xlsx(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "人物表"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    widths = [12, 12, 8, 8, 10, 18, 12, 18, 18, 12, 12, 16, 16, 16]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width

    table = Table(displayName="PeopleTable", ref=f"A1:N{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(table)

    summary = wb.create_sheet("分布摘要")
    summary.append(["维度", "类别", "人数", "占比"])
    for cell in summary[1]:
        cell.fill = fill
        cell.font = font

    for label, index in [("性别", 2), ("年龄段", 4), ("家庭状态", 5), ("职业大类", 6), ("户口类型", 9), ("户口省份", 10)]:
        counts = Counter(row[index] for row in rows)
        for category, count in sorted(counts.items(), key=lambda x: (-x[1], x[0])):
            summary.append([label, category, count, count / len(rows)])
    for cell in summary["D"][1:]:
        cell.number_format = "0.0%"

    params = wb.create_sheet("生成参数")
    params.append(["参数", "设定"])
    params.append(["样本人数", len(rows)])
    params.append(["随机种子", SEED])
    params.append(["性别", "男52% / 女48%"])
    params.append(["年龄段", "18-24 15%; 25-29 20%; 30-34 23%; 35-39 19%; 40-44 14%; 45-49 9%"])
    params.append(["户口类型", "上海户籍45% / 外省市户籍55%"])
    params.append(["家庭状态", "按年龄段条件概率生成"])
    params.append(["职业", "按年龄段条件概率生成"])
    params.append(["具体职位", "按职业生成；上班族职位层级随年龄调整"])
    params.append(["地点字段", "home_place_id / work_place_id / school_place_id 暂时留空"])
    params.append(["说明", "纯合成人口仿真，不代表真实个人"])
    for cell in params[1]:
        cell.fill = fill
        cell.font = font

    wb.save(path)


def main():
    out_dir = Path(__file__).resolve().parent
    rows = generate_rows()
    csv_path = out_dir / "shanghai_synthetic_population_10000.csv"
    xlsx_path = out_dir / "shanghai_synthetic_population_10000.xlsx"
    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)
    print(f"Generated {len(rows)} people")
    print(csv_path)
    print(xlsx_path)


if __name__ == "__main__":
    main()
